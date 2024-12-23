import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from openpyxl.styles import PatternFill
from time import sleep

# Carrega planilha
planilha_cadastro = openpyxl.load_workbook('cadastro_produto.xlsx')
sheet_cadastro = planilha_cadastro['ResultadoConsulta']

# Abre sistema no navegador
# Configure as opções do Chrome para iniciar em tela cheia
chrome_options = Options()
chrome_options.add_argument("--start-maximized")

wb = webdriver.Chrome(options=chrome_options)

wb.get('https://escritec.osayk.com.br/#/login')
sleep(5)

user_escritec = wb.find_element(By.NAME, 'username')
pass_escritec = wb.find_element(By.NAME, "password")

user_escritec.send_keys('***@gmail.com')
pass_escritec.send_keys('***')

# Login no Sistema
login_button = wb.find_element(
    By.XPATH, "//button[@class='btn btn-primary btn-block btn-flat']")
login_button.click()
sleep(12)

# Cadastro de dados
ctrl_estoque_button = wb.find_element(
    By.XPATH, "//a[@ng-href='#/controle-de-estoque']")
ctrl_estoque_button.click()
sleep(10)

cadastrar_produto_button = wb.find_element(
    By.XPATH, "//a[@class='btn-sm btn-primary']")
cadastrar_produto_button.click()
sleep(1)

linhas = sheet_cadastro.iter_rows(min_row=2, values_only=True)
linha_atual = 2

# Define o preenchimento verde
green_fill = PatternFill(start_color="6b8e23",
                         end_color="6b8e23", fill_type="solid")

for linha in linhas:
    id_produto, identificacao, descricao_produto, unidade, vlr_venda, cod_barra, ncm, cfop, \
        icmorigin, sittribicms, icms, sittribipi, ipi, sittribpis, alqpis, sittribcofins, \
        alqcofins, cest, indescrelev, cnpjfab, codbenef, percentualfcp, percentualst, \
        codanp, descricaoanp, regcodif, percglptroleo, percglpnat, percglpimp, valorpartida, \
        datacriacao, datatual, taxgas, displaymsg, ncmold = linha

    campo_codigo_produto = wb.find_element(By.NAME, 'code')
    campo_codigo_produto.send_keys(id_produto)

    campo_nome_produto = wb.find_element(By.NAME, 'name')
    campo_nome_produto.send_keys(descricao_produto)

    campo_descricao = wb.find_element(By.XPATH, "//input[@name='description']")
    campo_descricao.send_keys(descricao_produto)

    campo_valor = wb.find_element(
        By.XPATH, "//input[@name='averageUnitValue']")
    campo_valor.send_keys(vlr_venda)
    campo_valor.send_keys("00")

    campo_unidade_medida = wb.find_element(By.NAME, 'unityOfMeasure')
    select = Select(campo_unidade_medida)
    select.select_by_value(unidade)

    campo_ncm = wb.find_element(By.XPATH, "//input[@name='ncm']")
    campo_ncm.send_keys(ncm)

    campo_cfop = wb.find_element(By.XPATH, "//input[@name='cfop']")
    campo_cfop.send_keys("5102")

    campo_pis_cst = wb.find_element(By.NAME, 'pisCst')
    select = Select(campo_pis_cst)
    select.select_by_value("99")

    campo_icmscst = wb.find_element(By.XPATH, "//input[@name='icmsCst']")
    campo_icmscst.send_keys("0102")

    campo_cofins_cst = wb.find_element(By.NAME, 'cofinsCst')
    select = Select(campo_cofins_cst)
    select.select_by_value("99")

    campo_ipi_cst = wb.find_element(By.NAME, 'ipiCst')
    select = Select(campo_ipi_cst)
    select.select_by_value("99")

    salvar_continuar_button = wb.find_element(
        By.XPATH, "//button[@ng-click='saveAndContinue()']")
    salvar_continuar_button.click()
    sleep(1)

    for cell in sheet_cadastro[linha_atual]:
        cell.fill = green_fill

    linha_atual += 1

    planilha_cadastro.save('cadastro_produto.xlsx')

cancel_button = wb.find_element(By.XPATH, "//a[@ng-click='closeDialog()']")
cancel_button.click

wb.refresh()
